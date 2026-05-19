from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class RelatorioService:

    @staticmethod
    def exportar_impressoras_pdf(printers, filepath):
        """Gera PDF com lista de impressoras"""
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        # TÃ­tulo
        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#cba6f7'), spaceAfter=6)
        elements.append(Paragraph("RelatÃ³rio de Impressoras", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 15))

        # Resumo
        total = len(printers)
        operacionais = len([p for p in printers if p.status in ['Operacional', 'Em uso']])
        manutencao = len([p for p in printers if p.status in ['Em manutenÃ§Ã£o', 'ManutenÃ§Ã£o', 'Aguardando peÃ§a', 'Parada']])

        resumo_data = [
            ["Total Impressoras", str(total)],
            ["Operacionais", str(operacionais)],
            ["Em ManutenÃ§Ã£o/Paradas", str(manutencao)],
        ]
        t_resumo = Table(resumo_data, colWidths=[200, 100])
        t_resumo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#45475a')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cba6f7')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cba6f7')),
        ]))
        elements.append(t_resumo)
        elements.append(Spacer(1, 20))

        # Tabela principal
        header = ['PatrimÃ´nio', 'Modelo', 'Marca', 'Status', 'Local Atual', 'Ãšltima RevisÃ£o']
        data = [header]

        for p in printers:
            data.append([
                p.patrimonio,
                p.modelo,
                p.marca or '-',
                p.status,
                p.local_atual or '-',
                ''  # SerÃ¡ preenchido depois
            ])

        t = Table(data, colWidths=[55, 115, 70, 75, 120, 70])
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#cba6f7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e1e2e')),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#313244')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cba6f7')),
            ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#585b70')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#313244'), colors.HexColor('#45475a')]),
        ])
        t.setStyle(t_style)
        elements.append(t)

        # RodapÃ©
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Total de impressoras: {total} | Operacionais: {operacionais} | Em manutenÃ§Ã£o: {manutencao}", styles['Normal']))

        doc.build(elements)
        return filepath

    @staticmethod
    def exportar_impressoras_excel(printers, filepath):
        """Gera planilha Excel com lista de impressoras"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Impressoras"

        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='cba6f7', end_color='cba6f7', fill_type='solid')
        row_fill1 = PatternFill(start_color='313244', end_color='313244', fill_type='solid')
        row_fill2 = PatternFill(start_color='45475a', end_color='45475a', fill_type='solid')
        white_font = Font(name='Arial', size=10, color='FFFFFF')
        border = Border(
            left=Side(style='thin', color='cba6f7'),
            right=Side(style='thin', color='cba6f7'),
            top=Side(style='thin', color='cba6f7'),
            bottom=Side(style='thin', color='cba6f7')
        )

        # TÃ­tulo
        ws.merge_cells('A1:G1')
        ws['A1'] = f'RELATÃ“RIO DE IMPRESSORAS - {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='cba6f7')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Resumo
        total = len(printers)
        op = len([p for p in printers if p.status in ['Operacional', 'Em uso']])
        man = len([p for p in printers if p.status in ['Em manutenÃ§Ã£o', 'ManutenÃ§Ã£o', 'Aguardando peÃ§a', 'Parada']])

        ws['A3'] = f'Total: {total} | Operacionais: {op} | Em ManutenÃ§Ã£o: {man}'
        ws['A3'].font = Font(name='Arial', size=11, bold=True, color='89b4fa')

        # CabeÃ§alho
        headers = ['PatrimÃ´nio', 'Modelo', 'Serial', 'Marca', 'Status', 'Local Atual', 'ObservaÃ§Ã£o']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Dados
        for i, p in enumerate(printers, 6):
            dados = [p.patrimonio, p.modelo, p.serial or '-', p.marca or '-', p.status, p.local_atual or '-', p.observacao or '-']
            for col, val in enumerate(dados, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = white_font
                cell.fill = row_fill1 if i % 2 == 0 else row_fill2
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # Ajusta largura
        col_widths = [12, 22, 18, 15, 18, 25, 30]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        wb.save(filepath)
        return filepath

    @staticmethod
    def exportar_atividades_pdf(atividades, filepath):
        """Gera PDF com histÃ³rico de atividades"""
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=1*cm, rightMargin=1*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#cba6f7'))
        elements.append(Paragraph("RelatÃ³rio de Atividades / OS", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(atividades)}", styles['Normal']))
        elements.append(Spacer(1, 15))

        header = ['Data', 'PatrimÃ´nio', 'Tipo', 'DescriÃ§Ã£o', 'PeÃ§as', 'Origem', 'Destino', 'Status']
        data = [header]

        for a in atividades:
            data.append([
                a.event_at.strftime("%d/%m/%Y %H:%M") if a.event_at else '-',
                '',  # patrimÃ´nio (serÃ¡ preenchido)
                'ManutenÃ§Ã£o' if a.kind == 'MANUTENCAO' else 'MovimentaÃ§Ã£o',
                a.notes or '-',
                a.parts_used or '-',
                a.from_location or '-',
                a.to_location or '-',
                a.status_atividade or 'Concluida'
            ])

        t = Table(data, colWidths=[60, 40, 55, 120, 60, 55, 55, 50])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#cba6f7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e1e2e')),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#313244')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cba6f7')),
            ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#585b70')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#313244'), colors.HexColor('#45475a')]),
        ]))
        elements.append(t)

        doc.build(elements)
        return filepath

    @staticmethod
    def exportar_atividades_excel(atividades, filepath):
        """Gera Excel com atividades"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Atividades"

        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='cba6f7', end_color='cba6f7', fill_type='solid')
        white_font = Font(name='Arial', size=9, color='FFFFFF')
        fill1 = PatternFill(start_color='313244', end_color='313244', fill_type='solid')
        fill2 = PatternFill(start_color='45475a', end_color='45475a', fill_type='solid')
        border = Border(left=Side(style='thin', color='cba6f7'), right=Side(style='thin', color='cba6f7'),
                        top=Side(style='thin', color='cba6f7'), bottom=Side(style='thin', color='cba6f7'))

        ws.merge_cells('A1:H1')
        ws['A1'] = f'RELATÃ“RIO DE ATIVIDADES - {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='cba6f7')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = f'Total de atividades: {len(atividades)}'
        ws['A3'].font = Font(name='Arial', size=11, bold=True, color='89b4fa')

        headers = ['Data/Hora', 'Tipo', 'DescriÃ§Ã£o', 'PeÃ§as', 'Origem', 'Destino', 'Recibo', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for i, a in enumerate(atividades, 6):
            dados = [
                a.event_at.strftime("%d/%m/%Y %H:%M") if a.event_at else '-',
                'ManutenÃ§Ã£o' if a.kind == 'MANUTENCAO' else 'MovimentaÃ§Ã£o',
                a.notes or '-',
                a.parts_used or '-',
                a.from_location or '-',
                a.to_location or '-',
                a.numero_recibo or '-',
                a.status_atividade or 'Concluida'
            ]
            for col, val in enumerate(dados, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = white_font
                cell.fill = fill1 if i % 2 == 0 else fill2
                cell.border = border

        col_widths = [18, 15, 40, 20, 18, 18, 15, 15]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        wb.save(filepath)
        return filepath
